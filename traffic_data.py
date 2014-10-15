from openpyxl import Workbook
from utils import write_ws


class TrafficData():

    # DATA
    WS_SIMPLE_TBL_NAME = "principal"
    WS_DETAILS_TBL_NAME = "ver_detalle"
    FIELDS_SIMPLE_TBL = ["id_tramo", "Nro distrito", "Distrito",
                         "Limites del tramo", "Ini", "Fin", "TMDA", "Mas Info",
                         "Observaciones", "Link"]
    FIELDS_DETAILS_TBL = ["id_tramo", "id_tabla", "variable", "fila", "valor"]
    EXCEL_OUTPUT = "Datos de TMDA de la DNV - 2010 - corregido ruta 40.xlsx"

    def __init__(self):

        # create excel to save data
        self.wb = Workbook(optimized_write=True)

        # create sheet to store records of simple table
        self.ws_simple = self.wb.create_sheet(title=self.WS_SIMPLE_TBL_NAME)

        # create sheet to store records of details table
        self.ws_details = self.wb.create_sheet(title=self.WS_DETAILS_TBL_NAME)

    # PUBLIC
    def write_simple_record(self, record):
        dict_record = dict(zip(self.FIELDS_SIMPLE_TBL, record))
        write_ws(self.ws_simple, dict_record, self.FIELDS_SIMPLE_TBL)

    def write_details_record(self, record):
        dict_record = dict(zip(self.FIELDS_DETAILS_TBL, record))
        write_ws(self.ws_details, dict_record, self.FIELDS_DETAILS_TBL)

    def save(self, excel_output=None):
        excel_output = excel_output or self.EXCEL_OUTPUT
        self.wb.save(excel_output)
