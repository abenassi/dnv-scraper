# -*- coding: utf-8 -*-
import sys
import os
import time
from openpyxl import Workbook
from bs4 import BeautifulSoup
from urlparse import urljoin
from utils import (get_bs_from_static_site, extract_key_value_pairs_from_bs,
                   remove_accents, write_ws)

PARSER = "lxml"


class TrafficData():

    # DATA
    WS_SIMPLE_TBL_NAME = "tbl_simple"
    WS_DETAILS_TBL_NAME = "tbl_details"
    FIELDS_SIMPLE_TBL = ["id_tramo", "Nro distrito", "Distrito",
                         "Limites del tramo", "Ini", "Fin", "TMDA", "Mas Info",
                         "Observaciones", "Link"]
    FIELDS_DETAILS_TBL = ["id_tramo", "id_tabla", "variable", "fila", "valor"]
    EXCEL_OUTPUT = "Datos de TMDA de la DNV - 2010 - corregido ruta 40.xlsx"

    def __init__(self):

        # create excel to save data
        self.wb = Workbook(optimized_write=True)

        # create sheet to store records of simple table
        self.ws_simple = self.wb.create_sheet(title=self.WS_SIMPLE_TBL_NAME)

        # create sheet to store records of details table
        self.ws_details = self.wb.create_sheet(title=self.WS_DETAILS_TBL_NAME)

    # PUBLIC
    def write_simple_record(self, record):
        write_ws(self.ws_simple, record, self.FIELDS_SIMPLE_TBL)

    def write_details_record(self, record):
        write_ws(self.ws_details, record, self.FIELDS_DETAILS_TBL)

    def save(self, excel_output=None):
        excel_output = excel_output or self.EXCEL_OUTPUT
        self.wb.save(excel_output)


class RoadScraper():

    # DATA
    PARSER = "lxml"
    FIELDS_SIMPLE_TBL = ["id_tramo", "Nro distrito", "Distrito",
                         "Limites del tramo", "Ini", "Fin", "TMDA", "Mas Info",
                         "Observaciones", "Link"]
    FIELDS_DETAILS_TBL = ["id_tramo", "id_tabla", "variable", "fila", "valor"]

    def __init__(self, road_name, base_url, dict_rutas):

        # scrape parameters
        self.road_name = road_name
        self.base_url = base_url
        self.dict_rutas = dict_rutas

        # results
        self.tabla_ruta = []
        self.tabla_ruta_ver_detalle_dict = {}
        self.tabla_ruta_ver_detalle_db = []

    # PUBLIC
    def scrape(self):
        """
            toma la url de una ruta y devuelve 3 RVs en un arreglo:

            (1) un arreglo con la tabla con todos los tramos de la ruta con un
            codigo al principio identificador del tramo

            (2) un diccionario con todas las tablas "ver detalle" de aquellos tramos
            que no tienen promedio anual de la composicion pero si tienen un censo cobertura

            (3) un arreglo que almacena todos los datos "ver detalle" incluso de los tramos
            que no los tienen (con dicts vacios)
        """

        # parse base_url into a beautiful soup
        bs_road = get_bs_from_static_site(self.base_url, self.PARSER)

        # crea los arreglos y diccionarios que seran Return Values de la funcion
        tabla_ruta_ver_detalle = []
        tabla_ruta_censo_cobertura = {}

        id_section = 0
        # iterate rows
        for tr in bs_road.find_all("tr"):

            # get all elements of a row
            row_elements = tr.find_all("td", {"class": "FILA"})

            # check if row has 8 elements and proceed
            if len(row_elements) == 8:

                # new section has been found
                id_section += 1
                section_code = "{}_{}".format(self.road_name, id_section)
                print "Procesando tramo ", section_code

                # init a new row with section code
                row = [section_code]

                # iterate row elements adding each one to the row
                for element in row_elements:
                    row.append(element.get_text().strip())

                # procesa los datos de "Ver detalle"
                tablas_ver_detalle = {}
                if row_elements[6].get_text() == "ver detalle":

                    # agrega un campo más para el link del row_elements[6] (Ver detalle)
                    # cuando existe
                    link_part = row_elements[6].find_all("a")[0]["href"]
                    url_ver_detalle = urljoin(self.dict_rutas[self.road_name], link_part)
                    row.append(url_ver_detalle)

                    # extrae todas las tablas de "ver detalle"
                    tablas_ver_detalle = extract_tables_from_url_with_titles(
                        url_ver_detalle)

                else:
                    row.append("")

                # a la tabla_ruta le agrega la nueva fila de datos
                self.tabla_ruta.append(row)

                # agrega el diccionario de tablas extraidas de "ver detalle" con su
                # titulo al arreglo contenedor
                tabla_ruta_ver_detalle.append(tablas_ver_detalle)

        # transforma el diccionario de tablas de "ver detalle" en un arreglo tipo
        # tabla base de datos para excel
        self.tabla_ruta_ver_detalle_db = transformar_ver_detalle_dict_en_tabla(
            self.tabla_ruta, tabla_ruta_ver_detalle)

    def get_simple_records(self):

        for row in self.tabla_ruta:
            yield dict(zip(self.FIELDS_SIMPLE_TBL, row))

    def get_details_records(self):

        for row in self.tabla_ruta_ver_detalle_db:
            yield dict(zip(self.FIELDS_DETAILS_TBL, row))


# DATA
# scraping parameters
metodo = "GET"
base_url_part1 = "http://transito.vialidad.gov.ar:8080/SelCE_WEB/tmda_libro_web_"
base_url_part2 = "/index.html"

# other parameters
fields_simple_table = ["id_tramo", "Nro distrito", "Distrito",
                       "Limites del tramo", "Ini", "Fin", "TMDA",
                       "Mas Info", "Observaciones", "Link"]
fields_details_table = ["id_tramo", "id_tabla", "variable", "fila", "valor"]
fields_array = [fields_simple_table, fields_details_table]
years = [str(year) for year in list(xrange(2010, 2011))]


# METHODS
def extract_tables_from_url_with_titles(url):
    """
        Extrae todas las tablas de una url estatica que utilicen como tags "table", "thead", "tbody", "tr", "th" y "td"

        Devuelve un super diccionario donde el primer indice es la (1) tabla, el segundo indice es la (2) fila -donde
        la primer fila es el encabezado-, y el tercer indice es el (3) elemento de la fila.

        Los ids del diccionario son el primer texto que se encuentra en el html inmediatamente antes de la tabla!
    """

    # convierto en soup el url estatico
    soup = get_bs_from_static_site(url, PARSER)

    tablas = {}  # inicio un arreglo vacio para almacenar las tablas extraidas

    # itera entre las tablas de la soup
    for table in soup.find_all("table"):
        tabla = []  # nueva tabla vacia

        # entra en el encabezado de la tabla
        for thead in table.find_all("thead"):
            for tr in thead.find_all("tr"):
                row = []

                # toma todos los elementos de la fila del encabezado
                for th in tr.find_all("th"):
                    row.append(th.get_text())

                # almacena el encabezado en la tabla
                tabla.append(row)

        # entra en el contenido de la tabla
        for tbody in table.find_all("tbody"):

            # itera entre las filas del contenido de la tabla
            for tr in tbody.find_all("tr"):
                row = []

                # toma todos los elementos de la fila
                for td in tr.find_all("td"):
                    row.append(td.get_text())

                # almacena la fila en la tabla
                tabla.append(row)

        # toma el nombre de la tabla
        nombre_tabla = remove_accents(
            table.previous_sibling.previous_sibling.get_text())

        # agrega la tabla al super arreglo con todas las tablas
        tablas[nombre_tabla] = tabla

    return tablas


def scrape_link_rutas(base_url, parser):
    """ toma el url base de las rutas y devuelve un
    diccionario con sus links """

    # toma el url base y lo convierte en una bs
    soup_rutas = get_bs_from_static_site(base_url, parser)

    # extrae la tabla que contiene los links de las rutas
    pares_clave_valor_soup = soup_rutas.find_all("td", {"class": "FILA"})

    # para cada regitro de la tabla extrae el texto y el link
    dict_rutas = {}
    for element in pares_clave_valor_soup:
        dict_temp = extract_key_value_pairs_from_bs(element, "a", "href")
        dict_rutas[dict_temp[dict_temp.keys()[0]]] = dict_temp.keys()[0]

    # toma la lista de las rutas cuyos links fueron scrapeados
    lista_rutas = dict_rutas.keys()

    # convierte los link_parts de las rutas en links completos agregandole la
    # url base
    for ruta in lista_rutas:
        dict_rutas[ruta] = urljoin(base_url, dict_rutas[ruta])

    return dict_rutas


def transformar_ver_detalle_dict_en_tabla(tabla_ruta,
                                          tabla_ruta_ver_detalle_dict):

    # se crea el RV
    tabla_ruta_ver_detalle_db = []

    # itera con un indice por todos los tramos de la ruta
    for i in list(xrange(len(tabla_ruta_ver_detalle_dict))):

        # toma el id del tramo
        id_tramo = tabla_ruta[i][0].strip()

        # itera entre las tablas "ver detalle" de ese tramo si es que las hay
        if len(tabla_ruta_ver_detalle_dict[i]) > 0:
            for tabla in tabla_ruta_ver_detalle_dict[i]:

                # toma el id de la tabla
                id_tabla = tabla

                # itera entre las coordenadas fila-columna de la tabla
                for nro_fila in list(xrange(len(tabla_ruta_ver_detalle_dict[i][tabla]))):
                    for nro_col in list(xrange(len(tabla_ruta_ver_detalle_dict[i][tabla][nro_fila]))):

                        # si estamos en una fila con datos, no con encabezados
                        # incorporamos el valor
                        if nro_fila != 0:

                            # tomo la variable
                            variable = tabla_ruta_ver_detalle_dict[
                                i][tabla][0][nro_col].strip()

                            # tomo el nro de fila o la instancia
                            instancia = nro_fila

                            # tomo el valor
                            valor = tabla_ruta_ver_detalle_dict[i][
                                tabla][nro_fila][nro_col].strip()

                            # genero un nuevo registro con los datos tomados y
                            # lo agrego al arreglo
                            row = [
                                id_tramo, id_tabla, variable, instancia, valor]
                            tabla_ruta_ver_detalle_db.append(row)

    return tabla_ruta_ver_detalle_db


def main():

    # create a TrafficData object to store scraped data
    traffic_data = TrafficData()

    # llamado principal
    # itera por todos los años
    for anio in years:

        # forma el url base
        anio_base_url = base_url_part1 + anio + base_url_part2

        print "\n\nTomando datos del link base: ", anio_base_url
        print "Tomando datos del anio: ", anio, "\n\n"

        # toma los links de las rutas
        dict_rutas = scrape_link_rutas(anio_base_url, PARSER)
        #~ lista_rutas = dict_rutas.keys()
        lista_rutas = ["0040"]

        # itera por todas las rutas de ese año
        for ruta in lista_rutas:

            print "\nTomando los datos de la ruta: ", ruta, " en el anio: ",
            anio, "\n"

            # create scraper for road and scrape it
            road_scraper = RoadScraper(ruta, dict_rutas[ruta], dict_rutas)
            road_scraper.scrape()

            # write each simple record scraped to excel
            for record in road_scraper.get_simple_records():
                traffic_data.write_simple_record(record)

            # write each details record scraped to excel
            for record in road_scraper.get_details_records():
                traffic_data.write_details_record(record)

        traffic_data.save()



if __name__ == '__main__':
    main()
