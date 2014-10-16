import unittest
from dnv_scraper import scrape_traffic_data
from openpyxl import load_workbook


def compare_excels(excel1, excel2):
    """Compare two excels based on row iteration."""

    # load workbooks
    wb1 = load_workbook(excel1, use_iterators=True)
    wb2 = load_workbook(excel2, use_iterators=True)

    # check if sheets have same names
    if not wb1.get_sheet_names() == wb2.get_sheet_names():
        return False

    # iterate sheets
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):

        # iterate rows
        for row1, row2 in zip(ws1.iter_rows(), ws2.iter_rows()):

            # iterate cells
            for cell1, cell2 in zip(row1, row2):

                # evaluate cells value for equality
                if not cell1.value == cell2.value:
                    return False

    return True


class TestScrapeTrafficDataMethod(unittest.TestCase):

    def setUp(self):

        # names of excels to be compared
        self.test_compare_excel = "test_dnv_scraper_expected.xlsx"
        self.test_output_excel = "test_dnv_scraper.xlsx"

        # parameters to be passed to the tested function
        years = ["2010"]
        roads = ["0040"]

        # run scraper method
        scrape_traffic_data(years, roads, self.test_output_excel)

    def test_same_excels(self):
        self.assertTrue(compare_excels(self.test_compare_excel,
                                       self.test_output_excel))


def main():
    unittest.main()

if __name__ == '__main__':
    main()
