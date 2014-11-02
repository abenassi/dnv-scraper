import urllib2
from bs4 import BeautifulSoup
import unicodedata
from openpyxl import Workbook, load_workbook
from pprint import pprint


def compare_excels(excel1, excel2):
    """Compare two excels based on row iteration."""

    # load workbooks
    wb1 = load_workbook(excel1, use_iterators=True)
    wb2 = load_workbook(excel2, use_iterators=True)

    # check if sheets have same names
    if not wb1.get_sheet_names() == wb2.get_sheet_names():
        return False

    # iterate sheets
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):

        # iterate rows
        for row1, row2 in zip(ws1.iter_rows(), ws2.iter_rows()):

            # iterate cells
            for cell1, cell2 in zip(row1, row2):

                # evaluate cells value for equality
                if not cell1.value == cell2.value:
                    return False

    return True


def get_html_from_static_site(url):
    """Gets html from a static url, using urllib2."""

    headers_browser = {'User-Agent': 'Chrome/16.0.912.77'}
    request = urllib2.Request(url, '', headers=headers_browser)
    response = urllib2.urlopen(request)

    return response.read()


def get_bs_from_static_site(url, parser="html5lib"):
    """Gets beautiful soup from a static url."""

    html = get_html_from_static_site(url)
    soup = BeautifulSoup(html, parser)

    return soup


def extract_key_value_pairs_from_bs(key_value_pairs, main_tag, attribute):

    RV = {}
    key_value_list = key_value_pairs.find_all(main_tag)

    for record in key_value_list:
        RV[record[attribute].strip()] = record.get_text().strip()

    return RV


def remove_accents(data):
    return ''.join(x for x in unicodedata.normalize('NFKD', data)
                   if unicodedata.category(x)[0] == 'L').lower()


def write_tables_to_excel(tables, excel_output, fields=[]):
    """It gets a list of tables (a table being a list of lists) and a path for
    excel file, and returns excel file with tables in separated sheets."""

    # create new workbook
    wb = Workbook()

    # iterate through all tables
    num_table = 0
    for tabla in tables:

        # create new sheet for each table
        ws = wb.create_sheet()

        # iterate rows
        for i in xrange(len(tabla)):

            # iterate columns
            for j in xrange(len(tabla[i])):

                # copy cell content in the spreadsheet
                ws.cell(column=j + 1, row=i + 2).value = tabla[i][j]

        # if fields are passed, copy to first row
        if len(fields) > 0:

            for k in xrange(len(fields[num_table])):
                ws.cell(row=1, column=k + 1).value = fields[num_table][k]

        # increment num_table
        num_table += 1

    wb.save(excel_output)


def write_ws(ws, record, fields):
    """Add a record to a worksheet.

    Write fields first, if there are no records."""

    # check if worksheet is empty
    if ws.get_dimensions() == "A1":
        ws.append(fields)

    # add new row, with fields order
    new_row = []

    # extract data with field keys from record
    for field in fields:
        new_row.append(record[field])

    # add new row to worksheet
    ws.append(new_row)
